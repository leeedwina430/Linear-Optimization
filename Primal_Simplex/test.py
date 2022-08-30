
from main import *

#%% Example from book: Example 3.5
'''
A = np.array([[1,2,2,1,0,0],[2,1,2,0,1,0],[2,2,1,0,0,1]])
b = [20,20,20]
c = np.array([-10,-12,-12,0,0,0])

v1.TwoPhase_Simplex(A,b,c)
'''



#%% Example from book: Example 3.8
'''
A = np.array([[1,2,3,0],[-1,2,6,0],[0,4,9,0],[0,0,3,1]])
b = [3,2,5,1]
c = np.array([1,1,1,0])
v1.TwoPhase_Simplex(A,b,c)
'''
# With Gauss Elimination preprocess (A little performence gain in iteraions)
'''
v2.TwoPhase_Simplex(A,b,c)
'''



#%% Examplm from book: Exercise 3.17
'''
A = np.array([[1,3,0,4,1],[1,2,0,-1,1],[-1,-4,3,0,0]])
b = [2,2,1]
c = np.array([2,3,3,1,-2])

v1.TwoPhase_Simplex(A,b,c)
'''



#%% Performence Comparison
# 
# Randomly generate (m,n) size LP, compare the performance between
# implementation without and with Gauss Elimination
# 
# Note: This will take approximately [5 min]
'''
loop = 20
m,n = 300,400

iterations1 = []
times1 = []
for i in range(loop):
    print("seed=",i)

    np.random.seed(i)
    
    A = np.random.randn(m,n)*10
    b = list(np.random.randn(m)*10)
    c = np.random.randn(n)*10

    state,result,t,rtime = v1.TwoPhase_Simplex(A,b,c)
    iterations1.append(t)
    times1.append(rtime)
    
iterations2 = []
times2 = []
for i in range(loop):
    print("seed=",i)

    np.random.seed(i)
    A = np.random.randn(m,n)*10
    b = list(np.random.randn(m)*10)
    c = np.random.randn(n)*10

    state,result,t,rtime = v2.TwoPhase_Simplex(A,b,c)
    iterations2.append(t)
    times2.append(rtime)
    
print("TwoPhase_Simplex_01")
print("Average Iterations: ",np.mean(iterations1))
print("Average Running Time: ",np.mean(times1))

print("TwoPhase_Simplex_02")
print("Average Iterations: ",np.mean(iterations2))
print("Average Running Time: ",np.mean(times2))
'''



#%% Compare through Data Reading ([Solvable] Examples)
# Comparison between Version1, Version2 and Gurobi
# Note: All the data loaded from "data_filde_x.npz" are collected from random generation
'''
import gurobipy as gp
from gurobipy import GRB

As = np.load('data_file_A.npz')
bs = np.load('data_file_b.npz')
cs = np.load('data_file_c.npz')
Afiles = As.files
cfiles = cs.files
bfiles = bs.files

import gurobipy as gp
from gurobipy import GRB

m,n = As[Afiles[0]].shape

itersg = []
rtimesg = []

for i in range(len(Afiles)):
    try:

        model = gp.Model("redundency")

        x = model.addMVar(shape=n, name="x")
        model.setObjective(cs[cfiles[i]] @ x, GRB.MINIMIZE)
        model.addConstr(As[Afiles[i]] @ x == bs[bfiles[i]], name="e1")

        loA = np.identity(n)
        lo = np.zeros(n)
        model.addConstr(loA @ x >= lo, name="ine1")

        # Optimize model
        model.optimize()

        #print(x.X)
        #print('Obj: %g' % model.ObjVal)
        #print('Iter: %g' % model.IterCount)
        rtimesg.append(model.Runtime)
        itersg.append(model.IterCount)

    except gp.GurobiError as e:
        print('Error code ' + str(e.errno) + ": " + str(e))

    except AttributeError:
        print('Encountered an attribute error')

iters1,iters2 = [],[]
rtimes1,rtimes2 = [],[]

for i in range(len(Afiles)):
    state,opv,t,rtime = v1.TwoPhase_Simplex(As[Afiles[i]],bs[bfiles[i]],cs[cfiles[i]])
    iters1.append(t)
    rtimes1.append(rtime)

    state,opv,t,rtime = v1.TwoPhase_Simplex(As[Afiles[i]],bs[bfiles[i]],cs[cfiles[i]])
    iters2.append(t)
    rtimes2.append(rtime)


print("------------------------")
print("| Solvable Comparison: |")
print("------------------------")

print("TwoPhase_Simplex_01")
print("Average Iterations: ",np.mean(iters1))
print("Average Running Time: ",np.mean(rtimes1))
print()
print("TwoPhase_Simplex_02 (with GE)")
print("Average Iterations: ",np.mean(iters2))
print("Average Running Time: ",np.mean(rtimes2))
print()
print("Gurobi")
print("Average Iterations: ",np.mean(itersg))
print("Average Running Time: ",np.mean(rtimesg))
'''



#%% Example from the file "scrs8.mat"
# Note: It's not completely successful yet.
'''
from openpyxl.reader.excel import load_workbook
from TwoPhase_Simplex_02 import *

def read_xlsx(filename):
    wb = load_workbook(filename=filename)

    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])

    thelist = []
    if ws.max_column == 1:
        for i in range(1,ws.max_row+1):
            thelist.append(ws.cell(row=i, column=1).value)
    else:
        for i in range(1,ws.max_row+1):
            temp_list = []
            for j in range(1,ws.max_column+1):
                temp_list.append(ws.cell(row=i, column=j).value)
            thelist.append(temp_list)

    thearray = np.array(thelist)
    non = np.where(thearray==None)
    thearray[non] = 0
    thearray = thearray.astype(np.longfloat)
    return thearray

A = read_xlsx("A.xlsx")
b = read_xlsx("b.xlsx")
c = read_xlsx("c.xlsx")

TwoPhase_Simplex(A,b,c)
'''